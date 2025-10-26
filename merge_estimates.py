import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
import subprocess
import platform
import json


# -----------------------------
# Program Description
# -----------------------------
# This script consolidates multiple estimating spreadsheets (CSV/XLSX)
# into a single master spreadsheet (master_items.xlsx).
# Features:
# - Automatic category detection using categories.json
# - Standardizes column names and cleans Unit Cost
# - Removes duplicate items
# - Supports Overwrite and Append modes
# - Auto-resizes Excel columns
# - Opens master file automatically at the end
# - Logs processing info for each file and a summary

# -----------------------------
# Helper Functions
# -----------------------------
def auto_adjust_column_width(file_path):
    """
    Auto-adjust the width of all columns in an Excel file to fit content.

    Parameters:
        file_path (str): Path to the Excel file.
    """
    wb = load_workbook(file_path)
    ws = wb.active

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        adjusted_width = length + 2
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

    wb.save(file_path)
    print(f"✅ Auto-adjusted column widths in {file_path}")


def open_master_file(file_path):
    """
    Opens the master Excel file automatically in the default application.

    Parameters:
        file_path (str): Path to the Excel file.
    """
    try:
        if platform.system() == "Windows":
            os.startfile(file_path)
        elif platform.system() == "Darwin":  # macOS
            subprocess.call(["open", file_path])
        else:  # Linux
            subprocess.call(["xdg-open", file_path])
        print(f"✅ Opened master file: {file_path}")
    except Exception as e:
        print(f"⚠️ Could not open file: {e}")


def detect_category(filename, categories):
    """
    Detects the category of a spreadsheet based on its filename
    using a JSON mapping of keywords.

    Parameters:
        filename (str): The stem of the filename (without extension).
        categories (dict): Dictionary of category -> list of keywords.

    Returns:
        str: The detected category or "Unknown" if no match.
    """
    name = filename.lower()
    for category, keywords in categories.items():
        if any(keyword.lower() in name for keyword in keywords):
            return category
    return "Unknown"


# -----------------------------
# Data Cleaning Functions
# -----------------------------
def clean_dataframe(df, source_file, category):
    """
    Standardizes a dataframe and logs processing info:
    - Renames "Item Description" to "Description"
    - Cleans Unit Cost to float
    - Adds category and source file columns

    Parameters:
        df (pd.DataFrame): Input dataframe
        source_file (str): Name of the source file
        category (str): Category name

    Returns:
        pd.DataFrame: Cleaned dataframe or None if invalid
    """
    df.columns = df.columns.str.strip().str.lower()
    if "item description" in df.columns:
        df.rename(columns={"item description": "description"}, inplace=True)

    if "description" not in df.columns or "unit cost" not in df.columns:
        print(f"⚠️ Skipping {source_file} — missing required columns.")
        return None

    df["description"] = df["description"].astype(str).str.strip()
    df["unit cost"] = (
        df["unit cost"]
        .astype(str)
        .str.replace("[^0-9.]", "", regex=True)
        .replace("", "0")
        .astype(float)
    )

    df["category"] = category
    df["source file"] = source_file

    print(f"📄 Processed '{source_file}' | Items read: {len(df)} | Category: {category}")
    return df[["category", "description", "unit cost", "source file"]]


# -----------------------------
# Main Merge Function
# -----------------------------
def merge_estimates(input_folder="input_files", output_file="master_items.xlsx", mode="overwrite"):
    """
    Main function to merge multiple estimating spreadsheets into one master file.
    Supports both overwrite and append modes.

    Parameters:
        input_folder (str): Folder containing input files
        output_file (str): Output master Excel file
        mode (str): "overwrite" or "append"
    """
    input_path = Path(input_folder)
    input_path.mkdir(exist_ok=True)

    # Load categories JSON
    categories_file = Path("categories.json")
    if not categories_file.exists():
        print("⚠️ categories.json not found! Please create it in the script folder.")
        return
    with open(categories_file) as f:
        categories = json.load(f)

    # Collect all CSV/XLSX files
    all_files = [
        f for f in input_path.iterdir()
        if f.suffix.lower() in [".xlsx", ".csv"] and not f.name.startswith("~$")
    ]

    if not all_files:
        print("⚠️ No input files found in folder:", input_folder)
        return

    combined_data = []
    total_files_processed = 0
    total_items_processed = 0

    # -----------------------------
    # Process Each File
    # -----------------------------
    for file in all_files:
        category = detect_category(file.stem, categories)

        if file.suffix.lower() == ".csv":
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)

        df_clean = clean_dataframe(df, file.name, category)
        if df_clean is not None:
            combined_data.append(df_clean)
            total_files_processed += 1
            total_items_processed += len(df_clean)

    if not combined_data:
        print("⚠️ No valid data found in provided files.")
        return

    new_data = pd.concat(combined_data, ignore_index=True)

    # -----------------------------
    # Append or Overwrite Logic
    # -----------------------------
    if mode.lower() == "append" and Path(output_file).exists():
        existing_df = pd.read_excel(output_file)
        combined_df = pd.concat([existing_df, new_data], ignore_index=True)
        print("ℹ️ Append mode: Adding new unique items to existing master file.")
    else:
        combined_df = new_data
        if mode.lower() == "overwrite":
            print("ℹ️ Overwrite mode: Rebuilding master file from scratch.")

    # -----------------------------
    # Remove Duplicates and Log
    # -----------------------------
    before_dedup = len(combined_df)
    combined_df["normalized"] = (
        combined_df["description"].str.lower().str.replace(r"\s+", " ", regex=True)
    )
    combined_df.drop_duplicates(subset=["category", "normalized"], inplace=True)
    combined_df.drop(columns="normalized", inplace=True)
    after_dedup = len(combined_df)
    duplicates_removed = before_dedup - after_dedup

    print(f"\n🔹 Total files processed: {total_files_processed}")
    print(f"🔹 Total items before deduplication: {before_dedup}")
    print(f"🔹 Duplicates removed: {duplicates_removed}")
    print(f"🔹 Total unique items in master file: {after_dedup}\n")

    # -----------------------------
    # Save and Finalize
    # -----------------------------
    combined_df.to_excel(output_file, index=False)
    print(f"✅ Master spreadsheet saved: {output_file}")
    auto_adjust_column_width(output_file)
    open_master_file(output_file)


# -----------------------------
# Entry Point
# -----------------------------
if __name__ == "__main__":
    print("Select run mode:")
    print("1 = Overwrite (rebuild master file)")
    print("2 = Append (add only new unique items)")

    choice = input("Enter 1 or 2: ").strip()
    mode = "overwrite" if choice == "1" else "append"

    merge_estimates("input_files", "master_items.xlsx", mode=mode)
